import openpyxl, bs4, requests, re
import pandas as pd

#prerequisites: You must: 
#Have the relevant modules installed.
#Have an excel file adapted for the script (see #1 on line 30)
#Be looking for canadian stocks (in .to). Alternatively, you can modify script to allow searching other types of stocks. 

#Usefulness: Using excel functions, you can :
#Automatically calculate which stocks are the best to buy at a point in time.
#Update your porfolio value
#Estimate future dividend based on past dividend price.
#Have a template for adding new functions! Be creative! 


#Define variables to make the script a bit user friendly!
message = '''To use the script, type Stocks(Letter, Letter).
\nFirst letter is for Type(A=ETFs, B=Dividendes),
\nSecond letter is for Data(W=Price, X=Dividende; Y=Meta(Etfs Fees, past dividend); Z=All datas'''
print(message)
#Data type. You can add/modify data type, but you must also modify it in the Stocks function. 
A= 'ETFs'
B='Dividendes'
W= 'Price'
X='Dividende'
Y='Meta'
Z='All'

def Stocks(Type, Donnee):
#1 You must add some informations here.
    #This is the path to your file. 
    file = 'PATH\\FILE.xlsx'
    wb = openpyxl.load_workbook(file)
    #Col variable is the column number where you keep the acronyms for the stocks you want. Col A is 1, Col B is 2, etc. 
    Col='3'
    #Make sure you excel doc is setup up so Col is stock acronym, and that: 
    #Col+1 is Etf Fees (or blank otherwise), Col+2 is blank (or filled with other informations), Col+3 is price, Col+4 is dividend, Col+5 is dividend rate and Col+6 is past dividend.
    #Make sure you split individual stocks from etfs stocks. Their information are presented differently on Yahoo Finance. 

#2 This create a list according to the stocks you had written in column 'Col' and using the sheet the specified in function call (A, B).
    sheet = wb[Type]
    row = sheet.max_row
    Stocks_list = {}
    for cell in range(1, int(row)):
        read_stock = sheet.cell(row=cell, column=int(Col)).value
        Stocks_list[read_stock] = cell

#2.1 There might be some exceptions, like empty cells, titles or other types of datas. This is supposed to clear some exceptions. Adjust for your personnal exceptions if needed. 
    try: 
        Stocks_list.pop(None)
    except Exception as err:
        print('An exception happened: ' + str(err))        
    try:
        Stocks_list.pop('Cote')
    except Exception as err:
        print('An exception happened: ' + str(err))
    try:
        Stocks_list.pop('N/A')
    except Exception as err:
        print('An exception happened: ' + str(err))

#Here you have a dictionnary with datas in the form (Stock: row_number). This dict is printed. 
    print('Excel has been read. Here are the stocks'+str(Stocks_list))
    
#3 This is where the fun begins. 
#3.1. Each part will search for the relevant datas depending on data specified at first. They are ordered in numeral order of column according to the excel spreadsheat
#Etf Fees update. Will update if type is ETFs and if datas are 'Meta'. This is put in Col+1
    if Donnee == 'Meta' and Type=='ETFs':
        for k in Stocks_list.keys():
            url_link = f'https://finance.yahoo.com/quote/{k}.to'
            r = requests.get(url_link, headers = {'User-Agent':'Mozilla/5.0'})
            read_html_pandas_data = pd.read_html(r.text)

            try:
                Fees = float(read_html_pandas_data[1][1][6][0:4])/100
            except Exception as err:
                print('An exception happened: ' + str(err))
                continue
            cell_write = sheet.cell(row=Stocks_list[k], column=(int(Col)+1))
            cell_write.value = Fees
            print(str(Fees)+'is the new Fees value for stock '+str(k))


#This part will update stock price if 'Price' or 'all' data type were selected. This is written in col+3. 
    if Donnee == 'Price' or Donnee == 'All':
        for k in Stocks_list.keys():
            url_link = f'https://finance.yahoo.com/quote/{k}.to'
            r = requests.get(url_link, headers = {'User-Agent':'Mozilla/5.0'})
            read_html_pandas_data = pd.read_html(r.text)
            try:
                Close = float(read_html_pandas_data[0][1][0])
            except Exception as err:
                print('An exception happened: ' + str(err))
                continue

        #écrire dans le tableau maintenant
            cell_write = sheet.cell(row=Stocks_list[k], column=(int(Col)+3))
            cell_write.value = Close
            print(str(Close)+'is the new value for'+str(Donnee)+str(k))


#This will update dividend for individual stocks and write dividend in Col+4 and rate in Col+5. 
    if (Donnee == 'Dividende' or Donnee == 'All') and Type=='Dividendes': 
        for k in Stocks_list.keys():
            url_link = f'https://finance.yahoo.com/quote/{k}.TO/key-statistics?p={k}.TO'
            r = requests.get(url_link, headers = {'User-Agent':'Mozilla/5.0'})
            read_html_pandas_data = pd.read_html(r.text)
            Div = float(read_html_pandas_data[3][1][0])
            Div_Rate = float(read_html_pandas_data[3][1][1][0:4])/100

        #write dividend
            cell_write = sheet.cell(row=Stocks_list[k], column=(int(Col)+4))
            cell_write.value = Div
            print(str(Div)+'is the new dividend value for stock '+str(k))

        #write rate
            cell_write = sheet.cell(row=Stocks_list[k], column=(int(Col)+5))
            cell_write.value = Div_Rate
            print(str(Div_Rate)+'is the new dividend rate for stock '+str(k))
    
#This will update dividend rate for  ETFs. Data is written in Col+5. It's separate from past function becaue of how data is presented on Yahoo Finance (they dont give annual dividend for etfs, just percentage). 
    if (Donnee == 'Dividende' or Donnee == 'All') and Type=='ETFs': 
        for k in Stocks_list.keys():
            url_link = f'https://finance.yahoo.com/quote/{k}.TO/key-statistics?p={k}.TO'
            r = requests.get(url_link, headers = {'User-Agent':'Mozilla/5.0'})
            read_html_pandas_data = pd.read_html(r.text)

            try:
                Div = float(read_html_pandas_data[1][1][3][0:4])/100
            except Exception as err:
                print('An exception happened: ' + str(err))
                continue

        #write dividend
            cell_write = sheet.cell(row=Stocks_list[k], column=(int(Col)+5))
            cell_write.value = Div
            print(str(Div)+'is the new dividend rate for stock '+str(k))

#This will update past dividend using 2005 datas for the dividend. This data will be written in Col+6. If I remember correctly, it only works for stocks, not ETFs, because of how data or sorted on Yahoo Finance. 
    if Donnee == 'Meta' and Type=='Dividendes':
        for k in Stocks_list.keys():
            url_link = f'https://finance.yahoo.com/quote/{k}.to/history?period1=1104537600&period2=1135987200&interval=capitalGain|div|split&filter=div&frequency=1d&includeAdjustedClose=true'
            r = requests.get(url_link, headers = {'User-Agent':'Mozilla/5.0'})
            read_html_pandas_data = pd.read_html(r.text)
            TotalDiv = 0
            DivReg = re.compile(r'[0-9][.][0-9]+')
            try:
                AllDiv = DivReg.findall(str(read_html_pandas_data[0]))
            except:
                print('An exception happened: ' + str(err))
                continue
            
            for R in AllDiv:
                TotalDiv = TotalDiv + float(R)
            cell_write = sheet.cell(row=Stocks_list[k], column=(int(Col)+6))
            cell_write.value = TotalDiv
            print(str(TotalDiv)+'is the new Past Dividend value for stock '+str(k))

    #This will save the file at the end and reprint instruction message. 
    wb.save(file)
    print('information has been saved to'+str(file))
    print(message)

